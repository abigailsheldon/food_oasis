"""
USDA Food Directory Excel Import Script (Fixed)
Imports downloaded Excel files into Firebase Firestore.
"""

import re
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

# ============== CONFIGURATION ==============

FIREBASE_SERVICE_ACCOUNT_PATH = "food-oasis-fbab6-firebase-adminsdk-fbsvc-74a81df343.json"
STATE_FILTER = "Georgia"  # State name to filter (or None for all)

# Excel filenames 
EXCEL_FILES = {
    "Farmers Market": "farmers_market.xlsx",
    "Food Hub": "food_hub.xlsx",
    "On-Farm Market": "on_farm_market.xlsx",
    "CSA": "csa.xlsx",
}

# ============================================

# US States mapping
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY", "district of columbia": "DC"
}

# Reverse mapping (abbreviation to full name)
STATE_ABBREV_TO_NAME = {v: k.title() for k, v in US_STATES.items()}

# Initialize Firebase
cred = credentials.Certificate(FIREBASE_SERVICE_ACCOUNT_PATH)
firebase_admin.initialize_app(cred)
db = firestore.client()

# Counters
imported = 0
skipped = 0
failed = 0
filtered = 0


def extract_state_from_address(address):
    """Extract state name from a full address string."""
    if not address:
        return None
    
    address = str(address).strip()
    
    # Find state abbreviation pattern: ", XX " or ", XX" at end
    # Pattern: comma, space, 2 capital letters, then space or digit or end
    abbrev_match = re.search(r',\s*([A-Z]{2})(?:\s+\d|\s*$)', address)
    if abbrev_match:
        abbrev = abbrev_match.group(1)
        if abbrev in STATE_ABBREV_TO_NAME:
            return STATE_ABBREV_TO_NAME[abbrev]
    
    # Find full state name
    address_lower = address.lower()
    for state_name in US_STATES.keys():
        # Look for state name followed by space and zip or end
        if state_name in address_lower:
            return state_name.title()
    
    return None


def parse_hours(row):
    """Create default hours structure."""
    return {
        "Monday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Tuesday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Wednesday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Thursday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Friday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Saturday": {"isOpen": True, "open": "9:0", "close": "17:0"},
        "Sunday": {"isOpen": False, "open": None, "close": None},
    }


def safe_float(value):
    """Safely convert to float."""
    if value is None:
        return None
    try:
        return float(value)
    except:
        return None


def safe_str(value):
    """Safely convert to string."""
    if value is None:
        return ""
    return str(value).strip()


def import_row(row, headers, directory):
    """Import a single row into Firestore."""
    global imported, skipped, failed, filtered
    
    # Create dict from row using headers
    data = {}
    for i, header in enumerate(headers):
        if header and i < len(row):
            data[header.lower().strip()] = row[i]
    
    # Get address and extract state
    address = safe_str(data.get("location_address", ""))
    state = extract_state_from_address(address)
    
    # Filter by state if configured
    if STATE_FILTER and state and state.lower() != STATE_FILTER.lower():
        filtered += 1
        return  # Silently skip other states
    
    # Skip if no state found and filter is set
    if STATE_FILTER and not state:
        filtered += 1
        return
    
    name = safe_str(data.get("listing_name")) or safe_str(data.get("listing_desc")) or ""
    
    if not name:
        return
    
    # Get coordinates
    lat = safe_float(data.get("location_y"))
    lng = safe_float(data.get("location_x"))
    
    # Check if already exists
    existing = db.collection("businesses").where("name", "==", name).where("source", "==", "USDA").get()
    
    if len(existing) > 0:
        print(f"** {name[:50]} ** - already exists")
        skipped += 1
        return
    
    # Create business document
    try:
        doc_data = {
            "name": name,
            "address": address,
            "description": safe_str(data.get("listing_desc")) or "Local food vendor from USDA directory",
            "locationDetails": safe_str(data.get("location_desc")),
            "latitude": lat,
            "longitude": lng,
            "website": safe_str(data.get("media_website")),
            "phone": safe_str(data.get("contact_phone")),
            "email": safe_str(data.get("contact_email")),
            "directory": directory,
            "source": "USDA",
            "state": state,
            "acceptingReservations": False,
            "hours": parse_hours(data),
            "createdAt": firestore.SERVER_TIMESTAMP,
            "ownerUid": None,
            
            # Extra USDA data
            "certifications": safe_str(data.get("specialproductionmethods")),
            "paymentMethods": safe_str(data.get("acceptedpayment")),
            "foodAssistance": safe_str(data.get("fnap")) or safe_str(data.get("snap_option")),
            "products": safe_str(data.get("products")),
            "seasonality": safe_str(data.get("season_month")),
        }
        
        db.collection("businesses").add(doc_data)
        print(f"++ {name[:50]}")
        imported += 1
        
    except Exception as e:
        print(f"!! {name[:50]} - error: {e}")
        failed += 1


def import_excel_file(filepath, directory):
    """Import all rows from an Excel file."""
    global failed
    
    print(f"\n📂 Importing {directory} from {filepath}...")
    
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 2:
            print("   !! No data rows found")
            return
        
        headers = [str(h).lower().strip() if h else "" for h in rows[0]]
        print(f"   Found {len(rows) - 1} total rows")
        
        for row in rows[1:]:
            import_row(row, headers, directory)
        
        wb.close()
            
    except FileNotFoundError:
        print(f"   !!  File not found: {filepath} - skipping")
    except Exception as e:
        print(f"   !! Error reading file: {e}")
        failed += 1


def main():
    global imported, skipped, failed, filtered
    
    print("=" * 50)
    print("USDA Food Directory Excel Import")
    print("=" * 50)
    print(f"State filter: {STATE_FILTER or 'All states'}")
    print()
    
    for directory, filename in EXCEL_FILES.items():
        import_excel_file(filename, directory)
    
    print("\n" + "=" * 50)
    print("COMPLETE")
    print("=" * 50)
    print(f"+ Imported: {imported}")
    print(f"  Skipped (duplicates): {skipped}")
    print(f" Filtered (other states): {filtered}")
    print(f" Failed: {failed}")
    print()


if __name__ == "__main__":
    main()
